import openpyxl, csv
from io import TextIOWrapper
from decimal import InvalidOperation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from infrastructure.persistence.models import (
    CourseGroup, LaboratoryGroup, StudentEnrollment,
    AttendanceRecord, GradeRecord, Evaluation, Schedule, Course, CustomUser
)
from datetime import date, datetime, timedelta
from decimal import Decimal
from django.db import transaction
from django.db.models import Sum, F, ExpressionWrapper, DurationField, Count
from django.db.models.functions import Coalesce
from application.services.academic_calendar import get_group_sessions

# ====================================================================
# Funciones Auxiliares
# ====================================================================

def get_group_enrollments(group, group_type):
    """Función auxiliar para obtener matriculados ordenados por apellido/nombre."""
    if group_type == 'course':
        enrollments = StudentEnrollment.objects.filter(
            course=group.course,
            group=group,
            status='ACTIVO'
        ).select_related('student').order_by('student__last_name', 'student__first_name')
    else: # Lógica para LaboratoryGroup
        enrollments = StudentEnrollment.objects.filter(
            lab_assignment__lab_group=group,
            status='ACTIVO'
        ).select_related('student').order_by('student__last_name', 'student__first_name')
    return enrollments

# ====================================================================
# Vistas del Profesor
# ====================================================================

@login_required
def dashboard(request):
    """Dashboard principal del profesor"""
    
    if request.user.user_role != 'PROFESOR':
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('presentation:login')
    
    # Obtener grupos que enseña
    course_groups = CourseGroup.objects.filter(
        professor=request.user
    ).select_related('course')
    
    lab_groups = LaboratoryGroup.objects.filter(
        professor=request.user
    ).select_related('course')
    
    # Estadísticas rápidas
    total_courses = course_groups.count() + lab_groups.count()
    total_students = course_groups.aggregate(total=Sum('capacity'))['total'] or 0
    
    context = {
        'course_groups': course_groups,
        'lab_groups': lab_groups,
        'total_courses': total_courses,
        'total_students': total_students,
    }
    
    return render(request, 'professor/dashboard.html', context)


@login_required
def attendance(request):
    """Vista de asistencia del profesor"""
    
    if request.user.user_role != 'PROFESOR':
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('presentation:login')
    
    # Obtener grupos que enseña
    course_groups = CourseGroup.objects.filter(
        professor=request.user
    ).select_related('course')
    
    lab_groups = LaboratoryGroup.objects.filter(
        professor=request.user
    ).select_related('course')
    
    context = {
        'course_groups': course_groups,
        'lab_groups': lab_groups,
    }
    
    return render(request, 'professor/attendance.html', context)


@login_required
def record_attendance(request, group_id):
    """Registrar asistencia con validación de fecha y selector de sesión"""
    if request.user.user_role != 'PROFESOR':
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('presentation:login')
    
    # 1. Definición y Búsqueda de Grupo
    group = None
    group_type = None
    
    try:
        group = CourseGroup.objects.get(group_id=group_id, professor=request.user)
        group_type = 'course'
    except CourseGroup.DoesNotExist:
        try:
            group = LaboratoryGroup.objects.get(lab_id=group_id, professor=request.user)
            group_type = 'lab'
        except LaboratoryGroup.DoesNotExist:
            pass
            
    if not group:
        messages.error(request, 'Grupo no encontrado o no te pertenece.')
        return redirect('presentation:professor_attendance')

    # 2. Obtener Estudiantes Ordenados
    enrollments = get_group_enrollments(group, group_type)

    # 3. Calcular Sesiones del Semestre
    all_sessions = get_group_sessions(group) 
    current_session = None
    
    selected_date_str = request.GET.get('date')
    
    if selected_date_str:
        try:
            target_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
            current_session = next((s for s in all_sessions if s['date'] == target_date), None)
        except ValueError:
             messages.warning(request, 'Formato de fecha inválido.')
    
    if not current_session and all_sessions:
        # Si no hay selección, buscar la sesión de hoy o la última pasada
        today_session = next((s for s in all_sessions if s['is_today']), None)
        if today_session:
            current_session = today_session
        else:
            # Seleccionar la última sesión para ver historial por defecto
            current_session = all_sessions[-1] 

    # 4. Determinar si es editable
    # Solo es editable si es la fecha de hoy
    is_editable = current_session['is_today'] if current_session else False

    # 5. Obtener asistencias existentes para visualizarlas
    attendance_map = {}
    if current_session:
        existing_records = AttendanceRecord.objects.filter(
            enrollment__in=enrollments,
            session_number=current_session['number']
        )
        # Claves como cadena (UUID) para fácil manejo en templates
        attendance_map = {str(rec.enrollment_id): rec.status for rec in existing_records}
    
    if request.method == 'POST' and current_session:
        if not is_editable:
            messages.error(request, 'No se puede modificar la asistencia de fechas pasadas o futuras.')
            return redirect(f"{request.path}?date={current_session['date']}")

        session_number = current_session['number']
        session_date = current_session['date']

        for enrollment in enrollments:
            status = request.POST.get(f'status_{enrollment.enrollment_id}') 
            if status:
                AttendanceRecord.objects.update_or_create(
                    enrollment=enrollment,
                    session_number=session_number,
                    defaults={
                        'session_date': session_date,
                        'status': status,
                        'professor_ip': request.META.get('REMOTE_ADDR'),
                        'recorded_by': request.user
                    }
                )
                enrollment.calculate_attendance_percentage() 

        messages.success(request, f'Asistencia guardada para la sesión {session_number}.')
        return redirect(f"{request.path}?date={session_date}")

    context = {
        'group': group,
        'group_type': group_type,
        'enrollments': enrollments,
        'all_sessions': all_sessions,
        'current_session': current_session,
        'attendance_map': attendance_map,
        'is_editable': is_editable,
    }
    return render(request, 'professor/record_attendance.html', context)


@login_required
def attendance_report(request, group_id):
    """Vista de matriz de asistencia (Sábana de asistencia)"""
    if request.user.user_role != 'PROFESOR':
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('presentation:login')
        
    # 1. Búsqueda de Grupo 
    group = None
    group_type = None
    try:
        group = CourseGroup.objects.get(group_id=group_id, professor=request.user)
        group_type = 'course'
    except CourseGroup.DoesNotExist:
        try:
            group = LaboratoryGroup.objects.get(lab_id=group_id, professor=request.user)
            group_type = 'lab'
        except LaboratoryGroup.DoesNotExist:
            pass
            
    if not group:
        messages.error(request, 'Grupo no encontrado o no te pertenece.')
        return redirect('presentation:professor_attendance')
    
    # 2. Estudiantes y Registros
    enrollments = get_group_enrollments(group, group_type) 
    sessions = get_group_sessions(group)
    
    # Cargar todos los registros de asistencia con una sola consulta
    enrollment_ids = [e.enrollment_id for e in enrollments]
    records = AttendanceRecord.objects.filter(enrollment_id__in=enrollment_ids)
    
    # 3. Construir Matriz
    matrix = []
    
    # Mapear registros por enrollment_id (UUID) y session_number
    records_by_enrollment = {}
    for record in records:
        if record.enrollment_id not in records_by_enrollment:
            records_by_enrollment[record.enrollment_id] = {}
        records_by_enrollment[record.enrollment_id][record.session_number] = record.status

    for enrollment in enrollments:
        student_records = records_by_enrollment.get(enrollment.enrollment_id, {})
        
        row = {
            'student': enrollment.student,
            'enrollment': enrollment,
            'attendance_data': []
        }
        
        for sess in sessions:
            status = student_records.get(sess['number'], '-') 
            row['attendance_data'].append({
                'session_number': sess['number'],
                'status': status
            })
            
        matrix.append(row)

    if request.GET.get('export') == 'excel':
        # La función export_attendance_excel se define a continuación
        return export_attendance_excel(group, sessions, matrix)

    context = {
        'group': group,
        'sessions': sessions,
        'matrix': matrix
    }
    return render(request, 'professor/attendance_report.html', context)

def export_attendance_excel(group, sessions, matrix):
    """Función auxiliar para generar el Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencia"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    # Encabezado Superior
    ws['A1'] = f"Curso: {group.course.course_name}"
    ws['A2'] = f"Grupo: {group.group_code}"
    
    # Encabezados de Tabla (Fila 4)
    headers = ["Código", "Estudiante", "% Asist"]
    # Agregar fechas de sesiones como columnas
    for sess in sessions:
        headers.append(f"S{sess['number']}\n{sess['date'].strftime('%d/%m')}")
        
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True) # wrap_text para las fechas
        # Ajustar ancho para las sesiones
        if col_num > 3:
            ws.column_dimensions[get_column_letter(col_num)].width = 7 

    # Datos
    current_row = 5
    for row_data in matrix:
        ws.cell(row=current_row, column=1, value=row_data['student'].username)
        ws.cell(row=current_row, column=2, value=row_data['student'].get_full_name())
        ws.cell(row=current_row, column=3, value=f"{row_data['enrollment'].current_attendance_percentage}%")
        
        col_idx = 4
        for cell_data in row_data['attendance_data']:
            val = cell_data['status']
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            
            # Colores condicionales
            if val == 'F':
                cell.font = Font(color="FF0000", bold=True) # Rojo
            elif val == 'P':
                cell.font = Font(color="008000") # Verde
            elif val == 'J':
                 cell.font = Font(color="FFA500") # Naranja
                
            col_idx += 1
        current_row += 1

    # Ajustar anchos finales
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Asistencia_{group.group_code}.xlsx'
    wb.save(response)
    return response


@login_required
def grades(request):
    """Vista de notas del profesor"""
    if request.user.user_role != 'PROFESOR':
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('presentation:login')
    
    # Obtener cursos que enseña
    course_groups = CourseGroup.objects.filter(
        professor=request.user
    ).select_related('course').distinct()
    
    # Obtener cursos únicos
    courses = set([group.course for group in course_groups])
    
    context = {
        'courses': courses,
    }
    
    return render(request, 'professor/grades.html', context)


@login_required
def schedule(request):
    """Vista de horario del profesor"""
    
    if request.user.user_role != 'PROFESOR':
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('presentation:login')


    schedule_by_day = {
        'LUNES': [],
        'MARTES': [],
        'MIERCOLES': [],
        'JUEVES': [],
        'VIERNES': [],
    }

    total_duration = timedelta()
    total_classes = 0
    unique_course_codes = set()
    
    # 1. Obtener los horarios de Cursos
    course_schedules = Schedule.objects.filter(
        course_group__professor=request.user
    ).select_related(
        'course_group',        
        'course_group__course', 
        'room'                  
    )

    """
    # 2. Obtener los horarios de Laboratorio (Comentado en tu código original)
    # lab_schedules = LaboratorySchedule.objects.filter(
    #     lab_group__professor=request.user
    # ).select_related(
    #     'lab_group',           
    #     'lab_group__course',    
    #     'room'                  
    # )"""

    # 3. Procesar los horarios de Cursos
    for entry in course_schedules:
        time_start = datetime.combine(date.min, entry.start_time)
        time_end = datetime.combine(date.min, entry.end_time)
        duration = time_end - time_start
        
        total_duration += duration
        total_classes += 1
        unique_course_codes.add(entry.course_group.course.course_code)


        schedule_by_day[entry.day_of_week].append({
            'course': entry.course_group.course.course_name,
            'code': entry.course_group.course.course_code,
            'type': 'Teoría',
            'group': entry.course_group.group_code,
            'start_time': entry.start_time,
            'end_time': entry.end_time,
            'room': entry.room, 
        })
    
    """
    # 4. Procesar los horarios de Laboratorio (Comentado en tu código original)
    # for entry in lab_schedules:
    #     schedule_by_day[entry.day_of_week].append({
    #         'course': entry.lab_group.course.course_name,
    #         'code': entry.lab_group.course.course_code,
    #         'type': f'Lab {entry.lab_group.lab_nomenclature}',
    #         'group': entry.lab_group.lab_nomenclature,
    #         'start_time': entry.start_time,
    #         'end_time': entry.end_time,
    #         'room': entry.room, 
    #     }) """
    
    # 5. Ordenar por hora
    for day in schedule_by_day:
        schedule_by_day[day].sort(key=lambda x: x['start_time'])
    
    total_hours_decimal = total_duration.total_seconds() / 3600

    context = {
        'schedule_by_day': schedule_by_day,
        'total_horas': round(total_hours_decimal, 1),
        'total_clases': total_classes,
        'total_cursos': len(unique_course_codes),
    }
    
    return render(request, 'professor/schedule.html', context)

@login_required
def consolidated_grades(request, course_id):
    if request.user.user_role != 'PROFESOR':
        messages.error(request, 'No tienes permisos.')
        return redirect('presentation:login')

    # 1. Obtener Curso y validar profesor
    course = get_object_or_404(Course, course_id=course_id)
    if not CourseGroup.objects.filter(course=course, professor=request.user).exists():
        messages.error(request, 'No enseñas este curso.')
        return redirect('presentation:professor_dashboard')

    # 2. Estructura de Evaluaciones por Unidad
    all_evaluations = Evaluation.objects.filter(course=course).order_by('unit', 'evaluation_type')
    units_structure = {1: [], 2: [], 3: []}
    
    for evaluation in all_evaluations:
        if evaluation.unit in units_structure:
            units_structure[evaluation.unit].append(evaluation)

    # 3. Verificar Estado de Bloqueo por Unidad
    units_status = {}
    
    for unit_num, evals in units_structure.items():
        if not evals:
            units_status[unit_num] = True
            continue
        
        # Contar notas existentes para determinar si se bloquea
        grades_count = GradeRecord.objects.filter(
            evaluation__in=evals,
            enrollment__course=course
        ).values('enrollment').distinct().count()
        
        # Bloqueado si al menos hay notas registradas (regla de negocio actual)
        units_status[unit_num] = (grades_count > 0)

    # 4. Lógica de Guardado Manual (POST)
    if request.method == 'POST':
        unit_to_save = int(request.POST.get('unit_number'))
        
        # Validación: ¿Ya estaba bloqueada?
        if units_status.get(unit_to_save):
            messages.error(request, f'La Unidad {unit_to_save} ya fue registrada y no se puede modificar.')
            return redirect(request.path)

        evals_to_save = units_structure.get(unit_to_save, [])
        enrollments = StudentEnrollment.objects.filter(course=course, status='ACTIVO')

        try:
            with transaction.atomic():
                count_saved = 0
                errors = []
                
                for enrollment in enrollments:
                    for evaluation in evals_to_save:
                        input_name = f"grade_{enrollment.enrollment_id}_{evaluation.evaluation_id}"
                        raw_score = request.POST.get(input_name)

                        if raw_score and raw_score.strip():
                            try:
                                val = Decimal(raw_score)
                                if val < 0 or val > 20:
                                    errors.append(f"Nota inválida para {enrollment.student.get_full_name()}: {val}")
                                    continue
                                
                                GradeRecord.objects.update_or_create(
                                    enrollment=enrollment,
                                    evaluation=evaluation,
                                    defaults={
                                        'raw_score': val,
                                        'recorded_by': request.user
                                    }
                                )
                                count_saved += 1
                            except (ValueError, InvalidOperation):
                                errors.append(f"Formato inválido para {enrollment.student.get_full_name()}")
                    
                    # Recalcular promedio final del alumno
                    enrollment.calculate_final_grade()
                
                if errors:
                    messages.warning(request, f"Se guardaron {count_saved} notas. Errores: {'; '.join(errors[:3])}")
                else:
                    messages.success(request, f'Notas de la Unidad {unit_to_save} guardadas correctamente ({count_saved} registros).')
                
                return redirect(request.path)

        except Exception as e:
            messages.error(request, f'Error al guardar las notas: {str(e)}')

    # 5. Preparar Datos para la Tabla (GET)
    enrollments = StudentEnrollment.objects.filter(
        course=course, 
        status='ACTIVO'
    ).select_related('student').order_by('student__last_name', 'student__first_name')
    
    all_grades = GradeRecord.objects.filter(
        enrollment__in=enrollments, 
        evaluation__course=course
    ).select_related('evaluation')
    
    grades_map = {}
    
    for gr in all_grades:
        if gr.enrollment_id not in grades_map:
            grades_map[gr.enrollment_id] = {}
        grades_map[gr.enrollment_id][gr.evaluation_id] = gr.rounded_score

    matrix_data = []
    for enrollment in enrollments:
        row = {
            'enrollment': enrollment,
            'student': enrollment.student,
            'final_grade': enrollment.final_grade,
            'grades': {}
        }
        student_grades = grades_map.get(enrollment.enrollment_id, {})
        for unit_evals in units_structure.values():
            for ev in unit_evals:
                row['grades'][ev.evaluation_id] = student_grades.get(ev.evaluation_id)
        
        matrix_data.append(row)

    context = {
        'course': course,
        'units_structure': units_structure,
        'units_status': units_status,
        'matrix_data': matrix_data,
    }
    
    return render(request, 'professor/consolidated_grades.html', context)


@login_required
def upload_grades_csv(request, course_id):
    """Vista para cargar notas desde archivo CSV"""
    if request.user.user_role != 'PROFESOR':
        messages.error(request, 'No tienes permisos.')
        return redirect('presentation:login')
    
    course = get_object_or_404(Course, course_id=course_id)
    # Nombre corregido para redirección
    redirect_url_name = 'presentation:professor_record_grades'

    if not CourseGroup.objects.filter(course=course, professor=request.user).exists():
        messages.error(request, 'No enseñas este curso.')
        return redirect('presentation:professor_dashboard')
    
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        unit_number = int(request.POST.get('unit_number', 0))
        
        if not csv_file.name.endswith('.csv'):
            messages.error(request, 'El archivo debe ser .csv')
            return redirect(redirect_url_name, course_id=course_id)
        
        try:
            # Leer CSV
            file_data = TextIOWrapper(csv_file.file, encoding='utf-8')
            csv_reader = csv.DictReader(file_data)
            
            # Obtener evaluaciones de la unidad
            evaluations = Evaluation.objects.filter(
                course=course, 
                unit=unit_number
            ).order_by('evaluation_type')
            
            if evaluations.count() != 2:
                messages.error(request, f'La unidad {unit_number} debe tener exactamente 2 evaluaciones configuradas.')
                return redirect(redirect_url_name, course_id=course_id)
            
            eval_continua = evaluations.filter(evaluation_type='CONTINUA').first()
            eval_examen = evaluations.filter(evaluation_type='EXAMEN').first()
            
            if not eval_continua or not eval_examen:
                messages.error(request, 'Faltan evaluaciones CONTINUA o EXAMEN en la unidad.')
                return redirect(redirect_url_name, course_id=course_id)
            
            success_count = 0
            errors = []
            
            # Detectar nombres de columnas dinámicos o estáticos
            # Esto permite que el CSV generado anteriormente (continua1, examen1) funcione incluso si se sube a la unidad 2
            # Prioridad: Nombre dinámico (continua2) -> Nombre unidad 1 (continua1) -> Nombre genérico (continua)
            field_names = csv_reader.fieldnames
            
            col_continua = f'continua{unit_number}'
            col_examen = f'examen{unit_number}'

            # Fallback si no existen las columnas específicas de la unidad
            if col_continua not in field_names and 'continua1' in field_names:
                col_continua = 'continua1'
            if col_examen not in field_names and 'examen1' in field_names:
                col_examen = 'examen1'

            with transaction.atomic():
                for row in csv_reader:
                    cui = row.get('cui', '').strip()
                    # Usar las columnas detectadas
                    continua_score = row.get(col_continua, '').strip()
                    examen_score = row.get(col_examen, '').strip()
                    
                    if not cui:
                        continue
                    
                    try:
                        student = CustomUser.objects.get(username=cui, user_role='ALUMNO')
                        enrollment = StudentEnrollment.objects.get(
                            student=student,
                            course=course,
                            status='ACTIVO'
                        )
                        
                        # Guardar nota continua
                        if continua_score:
                            score_val = Decimal(continua_score)
                            if 0 <= score_val <= 20:
                                GradeRecord.objects.update_or_create(
                                    enrollment=enrollment,
                                    evaluation=eval_continua,
                                    defaults={
                                        'raw_score': score_val,
                                        'recorded_by': request.user
                                    }
                                )
                                success_count += 1
                        
                        # Guardar nota examen
                        if examen_score:
                            score_val = Decimal(examen_score)
                            if 0 <= score_val <= 20:
                                GradeRecord.objects.update_or_create(
                                    enrollment=enrollment,
                                    evaluation=eval_examen,
                                    defaults={
                                        'raw_score': score_val,
                                        'recorded_by': request.user
                                    }
                                )
                                success_count += 1
                        
                        # Recalcular promedio
                        enrollment.calculate_final_grade()
                        
                    except CustomUser.DoesNotExist:
                        errors.append(f"CUI {cui} no encontrado")
                    except StudentEnrollment.DoesNotExist:
                        errors.append(f"Estudiante {cui} no matriculado")
                    except (ValueError, InvalidOperation):
                        errors.append(f"Nota inválida para {cui}")
            
            if errors:
                messages.warning(request, f'Se cargaron {success_count} notas. Errores: {"; ".join(errors[:5])}')
            else:
                messages.success(request, f'Se cargaron {success_count} notas exitosamente.')
            
        except Exception as e:
            messages.error(request, f'Error al procesar CSV: {str(e)}')
        
        return redirect(redirect_url_name, course_id=course_id)
    
    return redirect(redirect_url_name, course_id=course_id)