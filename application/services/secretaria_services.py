import csv
import io
import json
import openpyxl
from datetime import time, datetime, timedelta
from typing import List, Dict, Optional
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from django.db import transaction
from django.db.models import Count, Q, Avg, Sum
from django.utils import timezone
from django.contrib.auth.hashers import make_password
from django.core.exceptions import ObjectDoesNotExist

from infrastructure.persistence.models import (
    Classroom,
    Course,
    CustomUser,
    CourseGroup,
    LabEnrollmentCampaign,
    StudentPostulation,
    LabAssignment,
    Schedule,
    StudentEnrollment,
    LaboratoryGroup,
    ExternalProfessor,
    Syllabus,
    SessionProgress,
    DAY_CHOICES,
)


class SecretariaService:
    """
    Servicio unificado para todas las operaciones de Secretaría:
    - Dashboard
    - Gestión de Horarios
    - Carga Masiva (CSV)
    - Reportes (Excel)
    - Gestión de Laboratorios
    """

    # ==================== DASHBOARD ====================
    @staticmethod
    def get_dashboard_stats():
        return {
            "classrooms_count": Classroom.objects.filter(is_active=True).count(),
            "courses_count": Course.objects.count(),
            "professors_count": CustomUser.objects.filter(
                user_role="PROFESOR", is_active=True
            ).count(),
            "students_count": CustomUser.objects.filter(
                user_role="ALUMNO", is_active=True
            ).count(),
        }

    # ==================== HORARIOS (SCHEDULE) ====================
    @staticmethod
    def get_schedule_context():
        """Prepara la data compleja para la vista de horarios"""
        groups_qs = (
            CourseGroup.objects.select_related("course", "professor")
            .prefetch_related("schedules", "schedules__room")
            .order_by("course__course_name", "group_code")
        )

        # Pesos para ordenamiento
        DAY_WEIGHTS = {
            "LUNES": 1,
            "MONDAY": 1,
            "1": 1,
            "MARTES": 2,
            "TUESDAY": 2,
            "2": 2,
            "MIERCOLES": 3,
            "WEDNESDAY": 3,
            "3": 3,
            "JUEVES": 4,
            "THURSDAY": 4,
            "4": 4,
            "VIERNES": 5,
            "FRIDAY": 5,
            "5": 5,
        }

        groups_data_list = []
        for group in groups_qs:
            schedules_list = list(group.schedules.all())
            # Lógica de ordenamiento
            schedules_list.sort(
                key=lambda s: (
                    DAY_WEIGHTS.get(str(s.day_of_week).strip().upper(), 99),
                    s.start_time,
                )
            )

            # JSON para el frontend
            schedules_json = [
                {
                    "day_of_week": schedule.day_of_week,
                    "start_time": schedule.start_time.strftime("%H:%M"),
                    "end_time": schedule.end_time.strftime("%H:%M"),
                    "room_id": (
                        str(schedule.room.classroom_id) if schedule.room else None
                    ),
                }
                for schedule in schedules_list
            ]

            groups_data_list.append(
                {
                    "group_object": group,
                    "sorted_schedules": schedules_list,
                    "schedules_json_data": schedules_json,
                }
            )

        return groups_data_list

    @staticmethod
    @transaction.atomic
    def update_course_group_schedules(group_id, horarios_data):
        """Actualiza los horarios de un grupo (Lógica de la API)"""
        group = CourseGroup.objects.get(group_id=group_id)
        dias_a_actualizar = set(h["day"] for h in horarios_data)

        # Limpiar horarios existentes en esos días
        if dias_a_actualizar:
            group.schedules.filter(day_of_week__in=dias_a_actualizar).delete()

        # Crear nuevos
        for h in horarios_data:
            Schedule.objects.create(
                course_group=group,
                room_id=h["room_id"],
                day_of_week=h["day"],
                start_time=h["start_time"],
                end_time=h["end_time"],
            )

    # ==================== CARGA MASIVA (CSV) ====================
    @staticmethod
    def process_student_csv(group_id, csv_file):
        """Procesa el CSV de alumnos: Crea usuarios, matricula y limpia"""
        try:
            group = CourseGroup.objects.get(group_id=group_id)
        except CourseGroup.DoesNotExist:
            return {"success": False, "error": "Grupo no encontrado"}

        # Leer archivo
        try:
            data = csv_file.read().decode("utf-8")
            io_string = io.StringIO(data)
            reader = csv.reader(io_string)
            next(reader, None)  # Saltar header
        except Exception as e:
            return {"success": False, "error": f"Error al leer archivo: {str(e)}"}

        student_cuis_in_csv = set()
        stats = {"created": 0, "enrolled": 0, "removed": 0}

        with transaction.atomic():
            # 1. Snapshot actual
            current_ids = set(
                StudentEnrollment.objects.filter(group=group).values_list(
                    "student__username", flat=True
                )
            )

            for row in reader:
                if not row or len(row) < 3:
                    continue

                cui = row[1].strip()
                full_name_raw = row[2].strip().upper()

                if not cui.isdigit():
                    continue
                student_cuis_in_csv.add(cui)

                # 2. Gestionar Usuario
                student = SecretariaService._get_or_create_student(
                    cui, full_name_raw, stats
                )

                # 3. Matricular
                if cui not in current_ids:
                    StudentEnrollment.objects.create(
                        student=student, group=group, course=group.course
                    )
                    stats["enrolled"] += 1

            # 4. Limpieza (Desmatricular los que ya no están en el CSV)
            to_remove = current_ids - student_cuis_in_csv
            if to_remove:
                deleted = StudentEnrollment.objects.filter(
                    group=group, student__username__in=to_remove
                ).delete()
                stats["removed"] = deleted[0]

            # 5. Actualizar metadatos del grupo
            group.students_loaded = True
            group.last_student_upload_at = timezone.now()
            group.capacity = len(student_cuis_in_csv)
            group.save()

        return {"success": True, "stats": stats, "group": group}

    @staticmethod
    def _get_or_create_student(cui, full_name_raw, stats):
        """Helper privado para lógica de usuario/email"""
        try:
            student = CustomUser.objects.get(username=cui)
            if not student.is_active:
                student.is_active = True
                student.account_status = "ACTIVO"
                student.save()
            return student
        except CustomUser.DoesNotExist:
            # Parsear nombres y generar email
            first, last, email = SecretariaService._parse_names_and_email(
                full_name_raw, cui
            )

            student = CustomUser.objects.create(
                username=cui,
                password=make_password(cui),
                email=email,
                first_name=first,
                last_name=last,
                user_role="ALUMNO",
                is_active=True,
                account_status="ACTIVO",
            )
            stats["created"] += 1
            return student

    @staticmethod
    def _parse_names_and_email(full_name_csv, cui):
        """Lógica pura de strings para nombres y emails"""
        try:
            parts = full_name_csv.replace('"', "").split(",")
            if len(parts) < 2:
                return "SinNombre", "SinApellido", f"err_{cui}@unsa.edu.pe"

            last_part = parts[0].strip().split("/")
            first_part = parts[1].strip()

            first_name = " ".join([n.capitalize() for n in first_part.split() if n])
            paterno = last_part[0].capitalize() if last_part else ""
            materno = last_part[1].capitalize() if len(last_part) > 1 else ""
            last_name = f"{paterno} {materno}".strip()

            # Email logic
            initial = first_name[0].lower() if first_name else "x"
            pat_slug = last_part[0].lower() if last_part else "x"

            email = f"{initial}{pat_slug}@unsa.edu.pe"

            # Verificación simple de colisión
            if CustomUser.objects.filter(email=email).exists():
                mat_slug = last_part[1].lower() if len(last_part) > 1 else ""
                email = f"{initial}{pat_slug}{mat_slug}@unsa.edu.pe"
                if CustomUser.objects.filter(email=email).exists():
                    email = f"{initial}{pat_slug}{cui}@unsa.edu.pe"

            return first_name, last_name, email
        except:
            return "Error", "Error", f"error_{cui}@unsa.edu.pe"

    # ==================== REPORTES (EXCEL) ====================
    @staticmethod
    def generate_grades_excel_workbook(group_id):
        """Genera el objeto Workbook con las notas"""
        group = CourseGroup.objects.get(group_id=group_id)
        course = group.course

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Notas {course.course_code}"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Encabezado Reporte
        ws["A1"] = f"REPORTE DE NOTAS - {course.course_name}"
        ws["A2"] = (
            f"Curso: {course.course_code} | Grupo: {group.group_code} | Créditos: {course.credits}"
        )

        # Columnas Dinámicas
        evals = course.evaluations.all().order_by("unit", "order")
        headers = ["CUI", "Alumno"] + [e.name for e in evals] + ["FINAL"]

        # Escribir Headers
        for col, txt in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=txt)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # Escribir Data
        enrollments = (
            StudentEnrollment.objects.filter(group=group, status="ACTIVO")
            .select_related("student")
            .prefetch_related("grade_records")
        )

        row_num = 5
        for env in enrollments:
            grades = {g.evaluation_id: g.rounded_score for g in env.grade_records.all()}

            ws.cell(row=row_num, column=1, value=env.student.username).border = (
                thin_border
            )
            ws.cell(row=row_num, column=2, value=env.student.get_full_name()).border = (
                thin_border
            )

            col_idx = 3
            for ev in evals:
                val = grades.get(ev.evaluation_id, "-")
                ws.cell(row=row_num, column=col_idx, value=val).border = thin_border
                col_idx += 1

            final = env.final_grade if env.final_grade is not None else "-"
            ws.cell(row=row_num, column=col_idx, value=final).border = thin_border
            row_num += 1

        return wb, f"Notas_{course.course_code}_{group.group_code}.xlsx"

    # ==================== GESTIÓN DE LABORATORIOS ====================

    @staticmethod
    def get_courses_with_lab():
        """Cursos que requieren laboratorio según sílabo"""
        return Course.objects.filter(
            syllabus__isnull=False, syllabus__lab_hours__gt=0
        ).distinct()

    @staticmethod
    def get_course_enrollment_count(course_id):
        return StudentEnrollment.objects.filter(
            course_id=course_id, status="ACTIVO"
        ).count()

    @staticmethod
    def calculate_lab_groups_needed(count):
        if count <= 30:
            return 1
        elif count <= 60:
            return 2
        return 3

    @staticmethod
    @transaction.atomic
    def create_lab_group(
        course_id,
        nomenclature,
        capacity,
        day,
        start,
        end,
        room_id=None,
        prof_id=None,
        ext_prof_data=None,
    ):
        """Crea un lab verificando conflictos"""
        result = {"success": False, "errors": []}
        try:
            # 1. Verificar Conflictos
            conflicts = SecretariaService.check_schedule_conflicts(
                course_id, day, start, end, room_id
            )
            if conflicts["has_conflict"]:
                result["errors"] = conflicts["messages"]
                return result

            # 2. Resolver Profesor
            ext_prof = None
            if ext_prof_data:
                ext_prof, _ = ExternalProfessor.objects.get_or_create(
                    full_name=ext_prof_data["full_name"],
                    defaults={
                        "email": ext_prof_data.get("email", ""),
                        "phone": ext_prof_data.get("phone", ""),
                    },
                )

            # 3. Crear
            LaboratoryGroup.objects.create(
                course_id=course_id,
                lab_nomenclature=nomenclature,
                capacity=capacity,
                day_of_week=day,
                start_time=start,
                end_time=end,
                room_id=room_id,
                professor_id=prof_id,
                external_professor=ext_prof,
            )
            result["success"] = True
        except Exception as e:
            result["errors"].append(str(e))
        return result

    @staticmethod
    def check_schedule_conflicts(
        course_id, day, start, end, room_id=None, exclude_lab_id=None
    ):
        """Verifica cruces de horario (Teoría, Sala, Otros Labs)"""
        msgs = []
        has_conflict = False

        # Helper interno
        def overlap(s1, e1, s2, e2):
            return s1 < e2 and s2 < e1

        # 1. Teoría
        theory_qs = Schedule.objects.filter(
            course_group__course_id=course_id, day_of_week=day
        )
        for s in theory_qs:
            if overlap(start, end, s.start_time, s.end_time):
                has_conflict = True
                msgs.append(f"Cruce con Teoría: {s.start_time}-{s.end_time}")

        # 2. Sala (Room)
        if room_id:
            # Cruce con clases normales
            occupied = Schedule.objects.filter(room_id=room_id, day_of_week=day)
            for s in occupied:
                if overlap(start, end, s.start_time, s.end_time):
                    has_conflict = True
                    msgs.append(f"Salón ocupado por clase: {s.start_time}")

            # Cruce con otros labs
            labs = LaboratoryGroup.objects.filter(room_id=room_id, day_of_week=day)
            if exclude_lab_id:
                labs = labs.exclude(lab_id=exclude_lab_id)
            for l in labs:
                if overlap(start, end, l.start_time, l.end_time):
                    has_conflict = True
                    msgs.append(f"Salón ocupado por Lab {l.lab_nomenclature}")

        return {"has_conflict": has_conflict, "messages": msgs}

    @staticmethod
    def get_available_classrooms(day, start, end):
        """Busca labs libres en ese horario"""

        def overlap(s1, e1, s2, e2):
            return s1 < e2 and s2 < e1

        occupied_ids = set()
        # Clases
        for s in Schedule.objects.filter(day_of_week=day).select_related("room"):
            if s.room and overlap(start, end, s.start_time, s.end_time):
                occupied_ids.add(s.room.classroom_id)
        # Labs
        for l in LaboratoryGroup.objects.filter(day_of_week=day).select_related("room"):
            if l.room and overlap(start, end, l.start_time, l.end_time):
                occupied_ids.add(l.room.classroom_id)

        return Classroom.objects.filter(
            is_active=True, classroom_type="LABORATORIO"
        ).exclude(classroom_id__in=occupied_ids)

    @staticmethod
    def delete_lab_group(lab_id):
        """
        Elimina un grupo de laboratorio validando reglas de negocio.
        Regla: No se puede eliminar si la campaña de inscripción ya cerró.
        """
        try:
            # 1. Obtener el lab
            from infrastructure.persistence.models import (
                LaboratoryGroup,
                LabEnrollmentCampaign,
            )

            lab = LaboratoryGroup.objects.get(lab_id=lab_id)

            # 2. Validar regla de negocio: Campaña cerrada
            campaign = LabEnrollmentCampaign.objects.filter(
                course=lab.course, is_closed=True
            ).first()

            if campaign:
                return {
                    "success": False,
                    "error": "No se puede eliminar el grupo. La inscripción del curso ya fue cerrada y procesada.",
                }

            # 3. Ejecutar eliminación
            lab.delete()
            return {
                "success": True,
                "message": "Grupo de laboratorio eliminado correctamente.",
            }

        except LaboratoryGroup.DoesNotExist:
            return {"success": False, "error": "El laboratorio no existe."}
        except Exception as e:
            return {"success": False, "error": f"Error al eliminar: {str(e)}"}

    # ==================== CAMPAÑA DE INSCRIPCIÓN ====================

    @staticmethod
    def can_enable_enrollment(course_id):
        """
        Paso 1: El Chequeo Previo.
        """
        total_students = SecretariaService.get_course_enrollment_count(course_id)

        # Sumamos la capacidad de todos los grupos (A, B, C...)
        total_capacity = (
            LaboratoryGroup.objects.filter(course_id=course_id).aggregate(
                total=Sum("capacity")
            )["total"]
            or 0
        )

        return {
            "can_enable": total_capacity >= total_students,
            "total_students": total_students,
            "total_capacity": total_capacity,
            "deficit": max(0, total_students - total_capacity),
        }

    @staticmethod
    @transaction.atomic
    def enable_lab_enrollment(course_id, days_duration=7):
        """
        Paso 2: ¡Abrir las puertas!
        Si todo está en orden, creamos la campaña. Esto permite que los botones
        de "Matricularse" aparezcan en la pantalla de los alumnos.
        """
        result = {"success": False, "errors": []}

        try:
            # Primero nos aseguramos que matemáticamente sea posible
            check = SecretariaService.can_enable_enrollment(course_id)
            if not check["can_enable"]:
                result["errors"].append(
                    f"Faltan cupos. Hay {check['total_students']} alumnos y solo {check['total_capacity']} espacios."
                )
                return result

            # Evitamos errores tontos: no crear una campaña si ya hay una corriendo
            active = LabEnrollmentCampaign.objects.filter(
                course_id=course_id, is_closed=False
            ).first()

            if active:
                result["errors"].append(
                    "Ya hay una inscripción activa para este curso."
                )
                return result

            # Creamos la campaña con fecha de inicio y fin
            now = timezone.now()
            campaign = LabEnrollmentCampaign.objects.create(
                course_id=course_id,
                start_date=now,
                end_date=now + timedelta(days=days_duration),
                is_closed=False,
            )

            result["success"] = True
            result["campaign_id"] = str(campaign.campaign_id)

        except Exception as e:
            result["errors"].append(str(e))

        return result

    @staticmethod
    def get_campaign_status(course_id):
        """
        El "Tablero de Control".
        Nos dice en tiempo real cómo va llenándose cada grupo.
        Calcula si un lab está vacío, normal, lleno o reventando de gente.
        """
        # Buscamos la última campaña creada
        campaign = (
            LabEnrollmentCampaign.objects.filter(course_id=course_id)
            .order_by("-created_at")
            .first()
        )

        if not campaign:
            return {"exists": False}

        # Contamos cuántos alumnos han pedido cada grupo
        labs = LaboratoryGroup.objects.filter(course_id=course_id).annotate(
            enrolled_count=Count(
                "postulations", filter=Q(postulations__campaign=campaign)
            )
        )

        labs_data = []
        for lab in labs:
            # Ponemos etiquetas visuales según qué tan lleno esté
            status = "empty"
            if lab.enrolled_count == 0:
                status = "empty"
            elif lab.enrolled_count < lab.capacity * 0.8:
                status = "normal"
            elif lab.enrolled_count <= lab.capacity:
                status = "almost-full"
            else:
                status = "exceeded"

            labs_data.append(
                {
                    "lab_id": str(lab.lab_id),
                    "nomenclature": lab.lab_nomenclature,
                    "capacity": lab.capacity,
                    "enrolled": lab.enrolled_count,
                    "available": max(0, lab.capacity - lab.enrolled_count),
                    "status": status,
                }
            )

        return {
            "exists": True,
            "campaign_id": str(campaign.campaign_id),
            "is_active": not campaign.is_closed,
            "is_closed": campaign.is_closed,
            "start_date": campaign.start_date.isoformat(),
            "end_date": campaign.end_date.isoformat(),
            "labs": labs_data,
        }

    @staticmethod
    def get_lab_enrolled_students(lab_id):
        """
        La lista de espera.
        Muestra quiénes se anotaron en un grupo específico y en qué orden llegaron.
        También avisamos si ese alumno tiene conflicto (para que el jefe sepa).
        """
        postulations = (
            StudentPostulation.objects.filter(lab_group_id=lab_id, status="PENDIENTE")
            .select_related("student")
            .order_by("timestamp")
        )

        students = []
        for i, post in enumerate(postulations, 1):
            students.append(
                {
                    "order": i,
                    "student_id": str(post.student.user_id),
                    "full_name": post.student.get_full_name(),
                    "email": post.student.email,
                    "timestamp": post.timestamp.strftime("%d/%m/%Y %H:%M"),
                    # Revisamos si el alumno tiene problemas de horario
                    "has_conflict": SecretariaService._check_student_schedule_conflict(
                        post.student.user_id, post.lab_group
                    ),
                }
            )

        return students

    @staticmethod
    def _check_student_schedule_conflict(student_id, lab_group):
        """
        El "Detector de Choques".
        Revisa la agenda del alumno para ver si tiene clases a la misma hora
        que el lab que quiere.
        Revisa:
        1. Sus clases de Teoría.
        2. Otros Laboratorios donde ya esté inscrito.
        """

        def overlap(s1, e1, s2, e2):
            return s1 < e2 and s2 < e1

        enrollments = StudentEnrollment.objects.filter(
            student_id=student_id, status="ACTIVO"
        ).select_related("course", "group")

        for enrollment in enrollments:
            if enrollment.group:
                schedules = Schedule.objects.filter(course_group=enrollment.group)
                for sch in schedules:
                    if sch.day_of_week == lab_group.day_of_week:
                        if overlap(
                            lab_group.start_time,
                            lab_group.end_time,
                            sch.start_time,
                            sch.end_time,
                        ):
                            return True

            if enrollment.lab_assignment:
                other_lab = enrollment.lab_assignment.lab_group
                if other_lab.lab_id != lab_group.lab_id:
                    if other_lab.day_of_week == lab_group.day_of_week:
                        if overlap(
                            lab_group.start_time,
                            lab_group.end_time,
                            other_lab.start_time,
                            other_lab.end_time,
                        ):
                            return True

        return False

    @staticmethod
    @transaction.atomic
    def close_lab_enrollment(course_id):
        """Solo cierra la campaña, los alumnos ya están matriculados."""
        result = {"success": False, "errors": []}

        try:
            campaign = LabEnrollmentCampaign.objects.filter(
                course_id=course_id, is_closed=False
            ).first()
            
            if not campaign:
                result["errors"].append("No hay campaña activa para cerrar.")
                return result

            # Solo cerrar la campaña
            campaign.is_closed = True
            campaign.closed_at = timezone.now()
            campaign.save()
            
            result["success"] = True

        except Exception as e:
            result["errors"].append(str(e))

        return result

    # Helper pequeño para no repetir código al guardar en base de datos
    @staticmethod
    def _assign_student(postulation, lab_group, method):
        LabAssignment.objects.create(
            postulation=postulation,
            student=postulation.student,
            lab_group=lab_group,
            assignment_method=method,
        )
        postulation.status = "ACEPTADO"
        postulation.save()

        # Actualizamos su matrícula oficial
        enrollment = StudentEnrollment.objects.filter(
            student=postulation.student, course_id=lab_group.course_id
        ).first()
        if enrollment:
            enrollment.lab_assignment = postulation.assignment
            enrollment.save()

    # ==================== ESTADÍSTICAS AVANZADAS ====================
    @staticmethod
    def get_statistics_context(semester_id=None):
        """
        Calcula todas las métricas para la vista de estadísticas.
        """
        # 1. Filtros Base
        enrollments_qs = StudentEnrollment.objects.filter(status="ACTIVO")
        courses_qs = Course.objects.all()
        groups_qs = CourseGroup.objects.all()

        if semester_id:
            courses_qs = courses_qs.filter(semester_id=semester_id)
            enrollments_qs = enrollments_qs.filter(course__semester_id=semester_id)
            groups_qs = groups_qs.filter(course__semester_id=semester_id)

        # 2. Análisis de Saturación
        groups_analysis = groups_qs.annotate(
            student_count=Count("enrollments", filter=Q(enrollments__status="ACTIVO"))
        )
        full, optimal, low = 0, 0, 0

        for group in groups_analysis:
            cap = group.capacity if (group.capacity and group.capacity > 0) else 40
            ratio = group.student_count / cap
            if ratio >= 0.9:
                full += 1
            elif ratio >= 0.5:
                optimal += 1
            else:
                low += 1

        saturation_json = json.dumps(
            {
                "labels": [
                    "Sobresaturados (>90%)",
                    "Ocupación Óptima",
                    "Baja Demanda (<50%)",
                ],
                "data": [full, optimal, low],
            }
        )

        # 3. Top 10 Cursos
        students_by_course = (
            enrollments_qs.values("course__course_code", "course__course_name")
            .annotate(count=Count("student"))
            .order_by("-count")[:10]
        )

        students_by_course_json = json.dumps(
            {
                "labels": [i["course__course_code"] for i in students_by_course],
                "names": [i["course__course_name"] for i in students_by_course],
                "data": [i["count"] for i in students_by_course],
            }
        )

        # 4. Carga Docente (Top 15)
        professors = CustomUser.objects.filter(
            user_role="PROFESOR", account_status="ACTIVO"
        )
        professors_load = []
        for prof in professors:
            # Calculamos horas únicas (evitando duplicados si un horario se repite en data sucia)
            schedules = Schedule.objects.filter(course_group__professor=prof).distinct()
            minutes = sum(
                [
                    (s.end_time.hour * 60 + s.end_time.minute)
                    - (s.start_time.hour * 60 + s.start_time.minute)
                    for s in schedules
                ]
            )
            hours = minutes / 60
            if hours > 0:
                professors_load.append(
                    {"name": prof.get_full_name(), "hours": round(hours, 1)}
                )

        professors_load.sort(key=lambda x: x["hours"], reverse=True)
        professors_load_json = json.dumps(
            {
                "labels": [p["name"] for p in professors_load[:15]],
                "data": [p["hours"] for p in professors_load[:15]],
            }
        )

        # 5. Avance de Sílabos
        syllabuses = Syllabus.objects.filter(course__in=courses_qs).select_related(
            "course"
        )
        syllabus_progress = []
        for syllabus in syllabuses:
            total = syllabus.sessions.count()
            completed = (
                SessionProgress.objects.filter(session__syllabus=syllabus).count()
                if total > 0
                else 0
            )
            pct = round((completed / total) * 100, 1) if total > 0 else 0
            syllabus_progress.append(
                {
                    "course_code": syllabus.course.course_code,
                    "course_name": syllabus.course.course_name,
                    "progress": pct,
                    "completed": completed,
                    "total": total,
                }
            )
        syllabus_progress.sort(key=lambda x: x["progress"])

        # 6. Uso de Aulas
        classrooms = (
            Classroom.objects.filter(is_active=True)
            .annotate(schedules_count=Count("schedules", distinct=True))
            .order_by("-schedules_count")
        )
        classrooms_usage = [
            {
                "code": c.code,
                "name": c.name,
                "usage_count": c.schedules_count,
                "capacity": c.capacity,
            }
            for c in classrooms
        ]

        # 7. KPIs Generales
        avg_size = (
            enrollments_qs.values("course")
            .annotate(c=Count("student"))
            .aggregate(a=Avg("c"))["a"]
        )

        return {
            "saturation_json": saturation_json,
            "students_by_course_json": students_by_course_json,
            "professors_load_json": professors_load_json,
            "syllabus_progress": syllabus_progress,
            "classrooms_usage": classrooms_usage,
            "total_students": enrollments_qs.values("student").distinct().count(),
            "total_courses": courses_qs.count(),
            "total_professors": professors.count(),
            "avg_class_size": round(avg_size or 0, 1),
        }

    # ==================== GESTIÓN DE RESERVAS ====================
    
    @staticmethod
    def get_pending_reservations():
        """
        Obtiene todas las reservas pendientes de aprobación.
        Ordenadas por fecha más cercana primero.
        """
        from infrastructure.persistence.models import ClassroomReservation
        
        return ClassroomReservation.objects.filter(
            status='PENDIENTE',
            reservation_date__gte=date.today()
        ).select_related(
            'classroom', 'professor'
        ).order_by('reservation_date', 'start_time')
    
    @staticmethod
    def get_all_reservations(status_filter=None):
        """
        Obtiene todas las reservas con filtro opcional por estado.
        """
        from infrastructure.persistence.models import ClassroomReservation
        
        qs = ClassroomReservation.objects.select_related(
            'classroom', 'professor', 'approved_by'
        ).order_by('-reservation_date', '-start_time')
        
        if status_filter:
            qs = qs.filter(status=status_filter)
        
        return qs
    
    @staticmethod
    @transaction.atomic
    def approve_reservation(reservation_id, secretaria_user):
        """
        Aprueba una reserva pendiente.
        """
        from infrastructure.persistence.models import ClassroomReservation
        
        try:
            reservation = ClassroomReservation.objects.get(
                reservation_id=reservation_id,
                status='PENDIENTE'
            )
            
            # Verificar disponibilidad nuevamente (por si acaso)
            conflicts = SecretariaService._check_reservation_conflicts(
                reservation.classroom_id,
                reservation.reservation_date,
                reservation.start_time,
                reservation.end_time,
                exclude_id=reservation_id
            )
            
            if conflicts:
                return {
                    'success': False,
                    'error': f'Conflicto detectado: {conflicts[0]}'
                }
            
            reservation.status = 'APROBADA'
            reservation.approved_by = secretaria_user
            reservation.approved_at = timezone.now()
            reservation.save()
            
            return {
                'success': True,
                'message': 'Reserva aprobada correctamente'
            }
            
        except ClassroomReservation.DoesNotExist:
            return {'success': False, 'error': 'Reserva no encontrada'}
    
    @staticmethod
    @transaction.atomic
    def reject_reservation(reservation_id, secretaria_user, reason=''):
        """
        Rechaza una reserva con motivo opcional.
        """
        from infrastructure.persistence.models import ClassroomReservation
        
        try:
            reservation = ClassroomReservation.objects.get(
                reservation_id=reservation_id,
                status='PENDIENTE'
            )
            
            reservation.status = 'RECHAZADA'
            reservation.approved_by = secretaria_user
            reservation.approved_at = timezone.now()
            reservation.rejection_reason = reason
            reservation.save()
            
            return {
                'success': True,
                'message': 'Reserva rechazada'
            }
            
        except ClassroomReservation.DoesNotExist:
            return {'success': False, 'error': 'Reserva no encontrada'}
    
    @staticmethod
    def _check_reservation_conflicts(classroom_id, date_obj, start_time, end_time, exclude_id=None):
        """
        Helper privado para detectar conflictos de una reserva.
        Retorna lista de mensajes de conflicto (vacía si no hay).
        """
        from infrastructure.persistence.models import ClassroomReservation
        
        def time_overlap(s1, e1, s2, e2):
            return s1 < e2 and s2 < e1
        
        conflicts = []
        
        # Verificar otras reservas aprobadas/pendientes
        other_reservations = ClassroomReservation.objects.filter(
            classroom_id=classroom_id,
            reservation_date=date_obj,
            status__in=['PENDIENTE', 'APROBADA']
        )
        
        if exclude_id:
            other_reservations = other_reservations.exclude(reservation_id=exclude_id)
        
        for res in other_reservations:
            if time_overlap(start_time, end_time, res.start_time, res.end_time):
                conflicts.append(
                    f"Conflicto con reserva de {res.professor.get_full_name()}"
                )
        
        return conflicts
