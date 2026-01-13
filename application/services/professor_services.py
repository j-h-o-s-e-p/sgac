import csv
import openpyxl
from io import TextIOWrapper
from datetime import datetime, date, timedelta, time
from decimal import Decimal, InvalidOperation
from django.db import transaction
from django.db.models import Avg, Max, Min, Count, Q
from django.core.files.storage import FileSystemStorage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Imports de Infraestructura y Dominio
from infrastructure.persistence.models import (
    CourseGroup,
    LaboratoryGroup,
    StudentEnrollment,
    AttendanceRecord,
    GradeRecord,
    Evaluation,
    Schedule,
    Course,
    CustomUser,
    Syllabus,
    SessionProgress,
    SyllabusSession,
    ClassroomReservation,
    Classroom,
)

# Imports de otros servicios
from application.services.academic_calendar import get_group_sessions, get_lab_sessions


class ProfessorService:

    # =========================================================================
    # 1. DASHBOARD Y GESTIÓN DE CURSOS
    # =========================================================================

    def get_professor_courses_cards(self, user):
        """
        Prepara la data para las tarjetas de 'Mis Cursos'.
        Unifica cursos de teoría y grupos de laboratorio en una sola lista.
        """
        course_groups = CourseGroup.objects.filter(professor=user).select_related(
            "course", "course__syllabus"
        )
        lab_groups = LaboratoryGroup.objects.filter(professor=user).select_related(
            "course", "course__syllabus"
        )

        cards = []

        # 1. Proceso grupos de teoría
        for group in course_groups:
            has_syllabus = (
                hasattr(group.course, "syllabus")
                and group.course.syllabus.syllabus_file
            )
            cards.append(
                {
                    "type": "TEORIA",
                    "group_obj": group,
                    "course": group.course,
                    "name": group.course.course_name,
                    "code": group.course.course_code,
                    "group_label": f"Grupo {group.group_code}",
                    "id_for_attendance": group.group_id,
                    "id_for_grades": group.group_id,
                    "has_syllabus": has_syllabus,
                    "syllabus_url": (
                        group.course.syllabus.syllabus_file.url
                        if has_syllabus
                        else None
                    ),
                }
            )

        # 2. Proceso grupos de laboratorio
        for lab in lab_groups:
            has_syllabus = (
                hasattr(lab.course, "syllabus") and lab.course.syllabus.syllabus_file
            )
            cards.append(
                {
                    "type": "LABORATORIO",
                    "group_obj": lab,
                    "course": lab.course,
                    "name": lab.course.course_name,
                    "code": lab.course.course_code,
                    "group_label": f"Lab {lab.lab_nomenclature}",
                    "id_for_attendance": lab.lab_id,
                    "id_for_grades": None,  # Labs generalmente no registran notas parciales aquí
                    "has_syllabus": has_syllabus,
                    "syllabus_url": (
                        lab.course.syllabus.syllabus_file.url if has_syllabus else None
                    ),
                }
            )

        return cards

    def get_dashboard_stats(self, user):
        """
        Calcula totales rápidos para el panel principal:
        - Estudiantes totales
        - Cursos activos
        - Progreso de sílabo por grupo
        """
        course_groups = (
            CourseGroup.objects.filter(professor=user)
            .select_related("course")
            .distinct()
        )
        lab_groups = (
            LaboratoryGroup.objects.filter(professor=user)
            .select_related("course")
            .distinct()
        )

        # Calcular estudiantes totales (evitando duplicados)
        total_students = (
            StudentEnrollment.objects.filter(group__in=course_groups, status="ACTIVO")
            .values("student")
            .distinct()
            .count()
        )

        total_students += (
            StudentEnrollment.objects.filter(
                lab_assignment__lab_group__in=lab_groups, status="ACTIVO"
            )
            .values("student")
            .distinct()
            .count()
        )

        # Agregar progreso de sílabo a cada grupo
        course_groups_with_progress = []
        for group in course_groups:
            progress = self._calculate_group_progress(group)
            course_groups_with_progress.append({"group": group, "progress": progress})

        return {
            "course_groups": course_groups_with_progress,
            "lab_groups": lab_groups,
            "total_courses": course_groups.count() + lab_groups.count(),
            "total_students": total_students,
        }

    def upload_syllabus(self, course_id, professor, pdf_file):
        """Sube el PDF del sílabo si el profesor tiene permiso"""
        course = Course.objects.get(course_id=course_id)

        # Validación de permiso: ¿Es profesor de teoría o lab de este curso?
        is_teoria = CourseGroup.objects.filter(
            course=course, professor=professor
        ).exists()
        is_lab = LaboratoryGroup.objects.filter(
            course=course, professor=professor
        ).exists()

        if not (is_teoria or is_lab):
            raise PermissionError("No tienes permiso para modificar este curso.")

        Syllabus.objects.update_or_create(
            course=course, defaults={"syllabus_file": pdf_file}
        )
        return course

    # =========================================================================
    # 2. GESTIÓN DE ASISTENCIA
    # =========================================================================

    def get_attendance_session_data(self, user, group_id, date_str=None):
        """
        Prepara todos los datos necesarios para la vista de tomar asistencia.
        """
        # 1. Buscar el grupo
        group, group_type = self._find_group_by_id(user, group_id)
        if not group:
            return {"error": "Grupo no encontrado o no tienes permiso."}

        # 2. Obtener alumnos
        enrollments = self._get_group_enrollments(group, group_type)

        # 3. Obtener calendario de sesiones
        if group_type == "course":
            all_sessions = get_group_sessions(group)
        else:
            all_sessions = get_lab_sessions(group)

        # 4. Determinar sesión actual
        current_session = self._determine_current_session(all_sessions, date_str)

        # 5. Variables de estado
        is_editable = False
        schedule_message = ""
        attendance_map = {}
        available_topics = []
        topics_stats = {"covered": 0, "quota": 0}

        if current_session:
            # Validar edición
            if current_session["is_today"]:
                if group_type == "lab":
                    is_editable, schedule_message = self._is_within_lab_schedule(group)
                else:
                    is_editable = True
                    schedule_message = "Sesión actual habilitada"
            else:
                schedule_message = "Solo lectura"

            # Cargar mapa de asistencia (inicia en 'F' si es hoy y está vacío)
            attendance_map = self._get_or_create_attendance_map(
                enrollments, current_session, user
            )

            # Si es teoría y es editable, cargar temas
            if is_editable and group_type == "course":
                available_topics = self.get_available_topics(group)
                covered = self.get_topics_covered_today_count(
                    group, current_session["date"]
                )
                topics_stats = {"covered": covered, "quota": max(0, 2 - covered)}

        return {
            "group": group,
            "group_type": group_type,
            "group_pk": group.group_id if group_type == "course" else group.lab_id,
            "enrollments": enrollments,
            "all_sessions": all_sessions,
            "current_session": current_session,
            "attendance_map": attendance_map,
            "available_topics": available_topics,
            "is_editable": is_editable,
            "schedule_message": schedule_message,
            "topics_covered_today": topics_stats["covered"],
            "topics_quota_left": topics_stats["quota"],
        }

    def save_attendance_process(
        self, user, group_id, date_obj, session_num, post_data, ip
    ):
        """Procesa el guardado de asistencia y temas"""
        group, group_type = self._find_group_by_id(user, group_id)
        if not group:
            return False

        enrollments = self._get_group_enrollments(group, group_type)

        self.save_attendance_and_topics(
            enrollments,
            session_num,
            date_obj,
            post_data,
            user,
            ip,
            group if group_type == "course" else None,
        )
        return True

    def get_attendance_report_matrix(self, user, group_id):
        """Genera la matriz para el reporte de asistencia"""
        group, group_type = self._find_group_by_id(user, group_id)
        if not group:
            return None

        enrollments = self._get_group_enrollments(group, group_type)
        sessions = (
            get_group_sessions(group)
            if group_type == "course"
            else get_lab_sessions(group)
        )

        # Obtener todos los registros en una sola consulta
        records = AttendanceRecord.objects.filter(enrollment__in=enrollments)
        records_map = {}
        for r in records:
            if r.enrollment_id not in records_map:
                records_map[r.enrollment_id] = {}
            records_map[r.enrollment_id][r.session_number] = r.status

        # Construir matriz
        matrix = []
        for env in enrollments:
            row = {"student": env.student, "enrollment": env, "attendance_data": []}
            for s in sessions:
                row["attendance_data"].append(
                    {
                        "session_number": s["number"],
                        "status": records_map.get(env.enrollment_id, {}).get(
                            s["number"], "-"
                        ),
                    }
                )
            matrix.append(row)

        return {
            "group": group,
            "group_id": group_id,
            "group_type": group_type,
            "sessions": sessions,
            "matrix": matrix,
        }

    def generate_attendance_excel(self, group, sessions, matrix):
        """Genera un archivo Excel con la matriz de asistencia"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Asistencia"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )

        # Encabezados Generales
        ws["A1"] = f"Curso: {group.course.course_name}"
        identifier = getattr(
            group, "group_code", getattr(group, "lab_nomenclature", "")
        )
        ws["A2"] = f"Grupo: {identifier}"

        # Encabezados de Tabla
        headers = ["Código", "Estudiante", "% Asist"]
        for sess in sessions:
            headers.append(f"S{sess['number']}\n{sess['date'].strftime('%d/%m')}")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            if col_num > 3:
                ws.column_dimensions[get_column_letter(col_num)].width = 6

        # Datos
        current_row = 5
        for row_data in matrix:
            ws.cell(row=current_row, column=1, value=row_data["student"].username)
            ws.cell(
                row=current_row, column=2, value=row_data["student"].get_full_name()
            )
            ws.cell(
                row=current_row,
                column=3,
                value=f"{row_data['enrollment'].current_attendance_percentage}%",
            )

            col_idx = 4
            for cell_data in row_data["attendance_data"]:
                val = cell_data["status"]
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                # Colores condicionales
                if val == "F":
                    cell.font = Font(color="FF0000", bold=True)
                elif val == "P":
                    cell.font = Font(color="008000")
                elif val == "J":
                    cell.font = Font(color="FFA500")
                col_idx += 1
            current_row += 1

        ws.column_dimensions["B"].width = 30
        return wb

    # =========================================================================
    # 3. GESTIÓN DE NOTAS
    # =========================================================================

    def get_grades_consolidation(self, user, group_id):
        """Prepara la matriz de notas por unidad"""
        try:
            group = CourseGroup.objects.select_related("course").get(
                group_id=group_id, professor=user
            )
        except CourseGroup.DoesNotExist:
            return None

        course = group.course

        # Estructura de evaluaciones
        evals = Evaluation.objects.filter(course=course).order_by(
            "unit", "evaluation_type"
        )
        units_structure = {1: [], 2: [], 3: []}
        for e in evals:
            if e.unit in units_structure:
                units_structure[e.unit].append(e)

        # Verificar si hay notas registradas
        units_status = {}
        for u, u_evals in units_structure.items():
            count = GradeRecord.objects.filter(
                evaluation__in=u_evals, enrollment__group=group
            ).count()
            units_status[u] = count > 0

        # Obtener alumnos y sus notas
        enrollments = (
            StudentEnrollment.objects.filter(group=group, status="ACTIVO")
            .select_related("student")
            .order_by("student__last_name")
        )

        all_grades = GradeRecord.objects.filter(
            enrollment__in=enrollments, evaluation__course=course
        )
        grades_map = {}
        for g in all_grades:
            if g.enrollment_id not in grades_map:
                grades_map[g.enrollment_id] = {}
            grades_map[g.enrollment_id][g.evaluation_id] = g.rounded_score

        # Armar matriz
        matrix_data = []
        for e in enrollments:
            row = {
                "enrollment": e,
                "student": e.student,
                "final_grade": e.final_grade,
                "grades": {},
            }
            s_grades = grades_map.get(e.enrollment_id, {})
            for u_evs in units_structure.values():
                for ev in u_evs:
                    row["grades"][ev.evaluation_id] = s_grades.get(ev.evaluation_id)
            matrix_data.append(row)

        return {
            "course": course,
            "group": group,
            "units_structure": units_structure,
            "units_status": units_status,
            "matrix_data": matrix_data,
        }


    def _parse_grade(self, raw, student_name, errors):
        """Parse and validate grade value"""
        if not raw or not raw.strip():
            return None

        try:
            val = Decimal(raw)
        except InvalidOperation:
            errors.append(f"Error formato: {student_name}")
            return None

        if not (0 <= val <= 20):
            errors.append(f"Nota fuera de rango: {student_name}")
            return None

        return val


    def save_grades_batch(self, course, unit_to_save, post_data, user):
        """Guarda notas masivamente desde formulario"""
        evaluations = Evaluation.objects.filter(course=course, unit=unit_to_save)
        enrollments = StudentEnrollment.objects.filter(course=course, status="ACTIVO")

        count = 0
        errors = []

        with transaction.atomic():
            for enrollment in enrollments:
                student_name = enrollment.student.get_full_name()

                for evaluation in evaluations:
                    input_name = f"grade_{enrollment.enrollment_id}_{evaluation.evaluation_id}"
                    raw = post_data.get(input_name)

                    val = self._parse_grade(raw, student_name, errors)
                    if val is None:
                        continue

                    GradeRecord.objects.update_or_create(
                        enrollment=enrollment,
                        evaluation=evaluation,
                        defaults={
                            "raw_score": val,
                            "recorded_by": user,
                            "is_locked": False,
                        },
                    )
                    count += 1

                enrollment.calculate_final_grade()

        return count, errors


    def _save_csv_grade(self, raw_value, enrollment, evaluation, user):
        """Validate and save a single grade from CSV"""
        if not raw_value:
            return False

        try:
            val = Decimal(raw_value)
        except InvalidOperation:
            return False

        if not (0 <= val <= 20):
            return False

        GradeRecord.objects.update_or_create(
            enrollment=enrollment,
            evaluation=evaluation,
            defaults={
                "raw_score": val,
                "recorded_by": user,
                "is_locked": False,
            },
        )
        return True


    def process_csv_grades(self, csv_file, course, unit_number, user):
        """Procesa carga masiva de notas desde CSV"""
        file_data = TextIOWrapper(csv_file.file, encoding="utf-8")
        csv_reader = csv.DictReader(file_data)

        evals = Evaluation.objects.filter(course=course, unit=unit_number)
        eval_cont = evals.filter(evaluation_type="CONTINUA").first()
        eval_exam = evals.filter(evaluation_type="EXAMEN").first()

        if not (eval_cont and eval_exam):
            raise ValueError("Faltan evaluaciones configuradas para esta unidad")

        col_c = f"continua{unit_number}"
        col_e = f"examen{unit_number}"

        success = 0
        errors = []

        with transaction.atomic():
            for row in csv_reader:
                cui = row.get("cui", "").strip()
                if not cui:
                    continue

                try:
                    student = CustomUser.objects.get(username=cui, user_role="ALUMNO")
                    enrollment = StudentEnrollment.objects.get(
                        student=student, course=course, status="ACTIVO"
                    )

                    for col, evaluation in ((col_c, eval_cont), (col_e, eval_exam)):
                        if self._save_csv_grade(
                            row.get(col), enrollment, evaluation, user
                        ):
                            success += 1

                    enrollment.calculate_final_grade()

                except Exception as e:
                    errors.append(f"CUI {cui}: {str(e)}")

        return success, errors


    # =========================================================================
    # 4. ESTADÍSTICAS Y REPORTES
    # =========================================================================

    def get_statistics_context(self, professor):
        """Calcula estadísticas detalladas para la vista de estadísticas"""
        groups = CourseGroup.objects.filter(professor=professor).select_related(
            "course"
        )
        if not groups.exists():
            return None

        all_enrollments = StudentEnrollment.objects.filter(
            group__in=groups, status="ACTIVO"
        )

        # Globales
        global_stats = all_enrollments.aggregate(
            promedio_general=Avg("final_grade"),
            asistencia_promedio=Avg("current_attendance_percentage"),
            total_alumnos=Count("enrollment_id"),
        )

        aprobados = all_enrollments.filter(final_grade__gte=10.5).count()
        desaprobados = all_enrollments.filter(final_grade__lt=10.5).count()
        habilitados = all_enrollments.filter(
            current_attendance_percentage__gte=70
        ).count()

        # Por Curso
        course_performance = []
        chart_labels = []
        chart_grades = []
        chart_attendance = []

        for group in groups:
            qs = group.enrollments.filter(status="ACTIVO")
            stats = qs.aggregate(
                avg=Avg("final_grade"),
                max=Max("final_grade"),
                min=Min("final_grade"),
                att=Avg("current_attendance_percentage"),
            )

            avg_grade = stats["avg"] or 0
            avg_att = stats["att"] or 0

            course_performance.append(
                {
                    "course_name": group.course.course_name,
                    "group_code": group.group_code,
                    "avg_grade": round(avg_grade, 2),
                    "max_grade": stats["max"] or 0,
                    "min_grade": stats["min"] or 0,
                    "avg_attendance": round(avg_att, 1),
                    "student_count": qs.count(),
                }
            )

            chart_labels.append(f"{group.course.course_name} - {group.group_code}")
            chart_grades.append(float(round(avg_grade, 2)))
            chart_attendance.append(float(round(avg_att, 1)))

        # Estudiantes en riesgo
        at_risk = (
            all_enrollments.filter(
                Q(final_grade__lt=10.5) | Q(current_attendance_percentage__lt=70)
            )
            .select_related("student", "course")
            .order_by("final_grade")[:10]
        )

        return {
            "no_data": False,
            "global_stats": global_stats,
            "aprobados_count": aprobados,
            "desaprobados_count": desaprobados,
            "habilitados_count": habilitados,
            "course_performance": course_performance,
            "at_risk_students": at_risk,
            "chart_labels": chart_labels,
            "chart_grades": chart_grades,
            "chart_attendance": chart_attendance,
        }

    def get_professor_schedule(self, professor):
        """Organiza el horario semanal (Teoría + Laboratorio)"""
        schedule_by_day = {
            "LUNES": [],
            "MARTES": [],
            "MIERCOLES": [],
            "JUEVES": [],
            "VIERNES": [],
        }
        total_duration = timedelta()
        total_classes = 0
        unique_course_codes = set()

        # 1. Teoría (Modelo Schedule)
        course_schedules = Schedule.objects.filter(
            course_group__professor=professor
        ).select_related("course_group", "course_group__course", "room")

        for entry in course_schedules:
            duration = datetime.combine(date.min, entry.end_time) - datetime.combine(
                date.min, entry.start_time
            )
            total_duration += duration
            total_classes += 1
            unique_course_codes.add(entry.course_group.course.course_code)

            if entry.day_of_week in schedule_by_day:
                schedule_by_day[entry.day_of_week].append(
                    {
                        "type": "Teoría",
                        "course": entry.course_group.course.course_name,
                        "code": entry.course_group.course.course_code,
                        "group": entry.course_group.group_code,
                        "start_time": entry.start_time,
                        "end_time": entry.end_time,
                        "room": entry.room if entry.room else "Sin Aula",
                        "_sort_key": entry.start_time,
                    }
                )

        # 2. Laboratorio (Modelo LaboratoryGroup)
        lab_groups = LaboratoryGroup.objects.filter(professor=professor).select_related(
            "course", "room"
        )

        for lab in lab_groups:
            duration = datetime.combine(date.min, lab.end_time) - datetime.combine(
                date.min, lab.start_time
            )
            total_duration += duration
            total_classes += 1
            unique_course_codes.add(lab.course.course_code)

            if lab.day_of_week in schedule_by_day:
                schedule_by_day[lab.day_of_week].append(
                    {
                        "type": "Laboratorio",
                        "course": lab.course.course_name,
                        "code": lab.course.course_code,
                        "group": lab.lab_nomenclature,
                        "start_time": lab.start_time,
                        "end_time": lab.end_time,
                        "room": lab.room if lab.room else "Lab",
                        "_sort_key": lab.start_time,
                    }
                )

        # Ordenar por hora
        final_schedule = {}
        for day, classes in schedule_by_day.items():
            if classes:
                classes.sort(key=lambda x: x["_sort_key"])
                final_schedule[day] = classes

        return {
            "schedule_by_day": final_schedule,
            "total_horas": round(total_duration.total_seconds() / 3600, 1),
            "total_clases": total_classes,
            "total_cursos": len(unique_course_codes),
        }

    def get_group_syllabus_details(self, group_id):
        """API para el timeline del sílabo"""
        try:
            group = CourseGroup.objects.get(group_id=group_id)
            syllabus = group.course.syllabus
        except:
            return {"error": "Grupo no encontrado o sin sílabo"}

        all_sessions = SyllabusSession.objects.filter(syllabus=syllabus).order_by(
            "session_number"
        )
        completed_ids = set(
            SessionProgress.objects.filter(course_group=group).values_list(
                "session_id", flat=True
            )
        )

        session_list = []
        for sess in all_sessions:
            session_list.append(
                {
                    "number": sess.session_number,
                    "topic": sess.topic,
                    "week": sess.week_number,
                    "is_completed": sess.session_id in completed_ids,
                }
            )

        return {
            "course_name": group.course.course_name,
            "group_name": f"Grupo {group.group_code}",
            "syllabus_data": [
                {"unit_name": "Desarrollo del Curso", "sessions": session_list}
            ],
        }

    # =========================================================================
    # 5. HELPERS PRIVADOS Y UTILITARIOS
    # =========================================================================

    def _find_group_by_id(self, user, group_id):
        """Busca en CourseGroup o LaboratoryGroup"""
        try:
            return CourseGroup.objects.get(group_id=group_id, professor=user), "course"
        except CourseGroup.DoesNotExist:
            try:
                return (
                    LaboratoryGroup.objects.get(lab_id=group_id, professor=user),
                    "lab",
                )
            except LaboratoryGroup.DoesNotExist:
                return None, None

    def _get_group_enrollments(self, group, group_type):
        """Obtiene alumnos ordenados por apellido"""
        if group_type == "course":
            return (
                StudentEnrollment.objects.filter(
                    course=group.course, group=group, status="ACTIVO"
                )
                .select_related("student")
                .order_by("student__last_name")
            )
        else:
            return (
                StudentEnrollment.objects.filter(
                    lab_assignment__lab_group=group, status="ACTIVO"
                )
                .select_related("student")
                .order_by("student__last_name")
            )

    def _determine_current_session(self, all_sessions, date_str):
        """Busca sesión por fecha o retorna la de hoy/última"""
        if date_str:
            try:
                target = datetime.strptime(date_str, "%Y-%m-%d").date()
                return next((s for s in all_sessions if s["date"] == target), None)
            except ValueError:
                pass

        today = next((s for s in all_sessions if s["is_today"]), None)

        last_session = all_sessions[-1] if all_sessions else None
        return today if today else last_session


    def _get_or_create_attendance_map(self, enrollments, session, user):
        """Obtiene asistencias existentes o crea 'F' por defecto si es hoy"""
        records = AttendanceRecord.objects.filter(
            enrollment__in=enrollments, session_number=session["number"]
        )

        if not records.exists() and session["is_today"]:
            new_recs = [
                AttendanceRecord(
                    enrollment=e,
                    session_number=session["number"],
                    session_date=session["date"],
                    status="F",
                    professor_ip="0.0.0.0",
                    recorded_by=user,
                )
                for e in enrollments
            ]
            AttendanceRecord.objects.bulk_create(new_recs)
            return {str(r.enrollment_id): "F" for r in new_recs}

        return {str(r.enrollment_id): r.status for r in records}

    def _calculate_group_progress(self, group):
        """Calcula porcentaje de avance del sílabo"""
        total = SyllabusSession.objects.filter(syllabus__course=group.course).count()
        if total == 0:
            return {"percentage": 0, "color": "secondary"}

        done = SessionProgress.objects.filter(course_group=group).count()
        pct = round((done / total) * 100, 1)
        return {"percentage": pct, "color": "success" if pct > 80 else "primary"}

    def _is_within_lab_schedule(self, lab):
        """Valida si la hora actual corresponde al horario del lab"""
        now = datetime.now()
        current_time = now.time()

        # Mapeo simple de días en inglés a español
        days = {
            "MONDAY": "LUNES",
            "TUESDAY": "MARTES",
            "WEDNESDAY": "MIERCOLES",
            "THURSDAY": "JUEVES",
            "FRIDAY": "VIERNES",
        }
        if days.get(now.strftime("%A").upper()) != lab.day_of_week:
            return False, f"El laboratorio es los {lab.get_day_of_week_display()}."

        # Margen de 15 minutos
        start_m = (
            datetime.combine(date.today(), lab.start_time) - timedelta(minutes=15)
        ).time()
        end_m = (
            datetime.combine(date.today(), lab.end_time) + timedelta(minutes=15)
        ).time()

        if start_m <= current_time <= end_m:
            return True, "Habilitado"
        return (
            False,
            f"Horario: {lab.start_time.strftime('%H:%M')} - {lab.end_time.strftime('%H:%M')}",
        )

    def save_attendance_and_topics(
        self, enrollments, session_num, session_date, post_data, user, ip, group
    ):
        """Guarda asistencia y temas (con límite diario)"""
        with transaction.atomic():
            # 1. Asistencia
            for enrollment in enrollments:
                status = post_data.get(f"attendance_{enrollment.enrollment_id}")
                if status:
                    AttendanceRecord.objects.update_or_create(
                        enrollment=enrollment,
                        session_number=session_num,
                        defaults={
                            "session_date": session_date,
                            "status": status,
                            "professor_ip": ip,
                            "recorded_by": user,
                        },
                    )
                    enrollment.calculate_attendance_percentage()

            # 2. Temas (Solo Teoría)
            if group:
                ids = post_data.getlist("topics_covered")
                if ids:
                    today_count = self.get_topics_covered_today_count(
                        group, session_date
                    )
                    limit = 2 - today_count
                    if limit > 0:
                        self._process_topics(ids[:limit], group, session_date, user)

    def _process_topics(self, topic_ids, group, date_obj, user):
        """Crea registros de SessionProgress sin duplicados"""
        for sid in topic_ids:
            if not SessionProgress.objects.filter(
                course_group=group, session_id=sid
            ).exists():
                SessionProgress.objects.create(
                    session_id=sid,
                    course_group=group,
                    completed_date=date_obj,
                    marked_by=user,
                )

    def get_available_topics(self, group):
        """Retorna temas no completados"""
        try:
            syllabus = group.course.syllabus
        except AttributeError:
            return []

        done = SessionProgress.objects.filter(course_group=group).values_list(
            "session_id", flat=True
        )

        return (
            SyllabusSession.objects.filter(syllabus=syllabus)
            .exclude(session_id__in=done)
            .order_by("session_number")
        )


    def get_topics_covered_today_count(self, group, date_obj):
        return SessionProgress.objects.filter(
            course_group=group, completed_date=date_obj
        ).count()

    # =========================================================================
    # 6. RESERVAS DE AULAS
    # =========================================================================

    def get_available_classrooms_for_reservation(self, date_obj, start_time, end_time):
        """
        Busca aulas disponibles para reserva en una fecha/hora específica.
        Considera: horarios regulares, labs, y otras reservas aprobadas.
        """

        def time_overlap(s1, e1, s2, e2):
            """Helper para detectar cruces de horario"""
            return s1 < e2 and s2 < e1

        # Obtener día de la semana
        day_name = date_obj.strftime("%A").upper()
        day_mapping = {
            "MONDAY": "LUNES",
            "TUESDAY": "MARTES",
            "WEDNESDAY": "MIERCOLES",
            "THURSDAY": "JUEVES",
            "FRIDAY": "VIERNES",
        }
        day_spanish = day_mapping.get(day_name, "LUNES")

        occupied_ids = set()

        # 1. Aulas ocupadas por horarios regulares de teoría
        regular_schedules = Schedule.objects.filter(
            day_of_week=day_spanish
        ).select_related("room")

        for schedule in regular_schedules:
            if schedule.room and time_overlap(
                start_time, end_time, schedule.start_time, schedule.end_time
            ):
                occupied_ids.add(schedule.room.classroom_id)

        # 2. Aulas ocupadas por laboratorios
        labs = LaboratoryGroup.objects.filter(day_of_week=day_spanish).select_related(
            "room"
        )

        for lab in labs:
            if lab.room and time_overlap(
                start_time, end_time, lab.start_time, lab.end_time
            ):
                occupied_ids.add(lab.room.classroom_id)

        # 3. Aulas con reservas aprobadas o pendientes en esa fecha
        reservations = ClassroomReservation.objects.filter(
            reservation_date=date_obj, status__in=["PENDIENTE", "APROBADA"]
        ).select_related("classroom")

        for reservation in reservations:
            if time_overlap(
                start_time, end_time, reservation.start_time, reservation.end_time
            ):
                occupied_ids.add(reservation.classroom.classroom_id)

        # Retornar aulas disponibles
        available = (
            Classroom.objects.filter(is_active=True)
            .exclude(classroom_id__in=occupied_ids)
            .order_by("classroom_type", "name")
        )

        return available

    def create_classroom_reservation(
        self, professor, classroom_id, date_obj, start_time, end_time, purpose
    ):
        """
        Crea una nueva reserva de aula (estado PENDIENTE).
        Valida:
        - Horario dentro de rango permitido (7am - 8:10pm)
        - Disponibilidad del aula
        - Duración mínima 1 hora
        """

        errors = []

        # Validación 1: Horario permitido
        MIN_TIME = time(7, 0)
        MAX_TIME = time(20, 10)

        if start_time < MIN_TIME or end_time > MAX_TIME:
            errors.append("El horario debe estar entre 7:00 AM y 8:10 PM")

        # Validación 2: Duración mínima
        duration = datetime.combine(date.min, end_time) - datetime.combine(
            date.min, start_time
        )
        if duration.total_seconds() < 3600:  # 1 hora mínimo
            errors.append("La reserva debe ser de al menos 1 hora")

        # Validación 3: Fecha no pasada
        if date_obj < date.today():
            errors.append("No puedes reservar fechas pasadas")

        # Validación 4: Disponibilidad
        if not errors:
            available = self.get_available_classrooms_for_reservation(
                date_obj, start_time, end_time
            )
            if not available.filter(classroom_id=classroom_id).exists():
                errors.append("El aula no está disponible en ese horario")

        if errors:
            return {"success": False, "errors": errors}

        # Crear reserva
        try:
            with transaction.atomic():
                reservation = ClassroomReservation.objects.create(
                    classroom_id=classroom_id,
                    professor=professor,
                    reservation_date=date_obj,
                    start_time=start_time,
                    end_time=end_time,
                    purpose=purpose,
                    status="PENDIENTE",
                )

                return {
                    "success": True,
                    "reservation_id": str(reservation.reservation_id),
                    "message": "Reserva creada. Pendiente de aprobación por Secretaría.",
                }
        except Exception as e:
            return {"success": False, "errors": [str(e)]}

    def get_professor_reservations(self, professor, include_history=False):
        """
        Lista las reservas del profesor.
        Si include_history=False, solo muestra activas (PENDIENTE/APROBADA).
        Si True, muestra todo el historial.
        """

        qs = ClassroomReservation.objects.filter(professor=professor).select_related(
            "classroom", "approved_by"
        )

        if not include_history:
            qs = qs.filter(
                status__in=["PENDIENTE", "APROBADA"], reservation_date__gte=date.today()
            )

        return qs.order_by("-reservation_date", "-start_time")

    def cancel_reservation(self, reservation_id, professor):
        """
        Cancela una reserva (solo si es del profesor y está activa).
        """

        try:
            reservation = ClassroomReservation.objects.get(
                reservation_id=reservation_id, professor=professor
            )

            if not reservation.can_cancel():
                return {
                    "success": False,
                    "error": "Esta reserva no puede ser cancelada",
                }

            reservation.status = "CANCELADA"
            reservation.save()

            return {"success": True, "message": "Reserva cancelada correctamente"}

        except ClassroomReservation.DoesNotExist:
            return {"success": False, "error": "Reserva no encontrada"}
